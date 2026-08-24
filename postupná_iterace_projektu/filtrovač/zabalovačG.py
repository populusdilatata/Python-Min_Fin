from __future__ import annotations

import logging
import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.cell.cell import Cell

# Architektura              9.5/10
# Čitelnost                 9.5/10
# Dokumentace               9.0/10
# Logging                   9.5/10
# Python styl               8.5/10
# Robustnost                9.0/10
# Udržovatelnost            9.5/10

# ======================================================
# KONFIGURACE
# ======================================================

CATEGORIES = (
    "ABSEN",
    "INDIV",
    "LEKAR",
    "DOVOL",
    "NEMOC",
    "DROBY_OCR",
    "DROBY_OTECD",
    "DROBY_PLACV",
    "DROBY_STUDV",
    "DROBY_SVZOZ",
)

SOURCE_ROOT = Path("porovnavač")
TARGET_ROOT = Path("PRO_Sko")

HELICOPTER_FOLDER = Path("helikoptérový pohled")

SELECTED_MONTHS = frozenset({"06"})

MONTH_COLUMN_INDEX = 1
THIRD_SHEET_INDEX = 2

FIRST_DATA_SHEET = 1
LAST_DATA_SHEET = 3
DATA_START_ROW = 2

# ======================================================
# BARVY
# ======================================================

YELLOW_FONT = Font(color="FFFF00")

BLACK_FILL = PatternFill(fill_type="solid", start_color="000000", end_color="000000",)

GREEN_FILL = PatternFill(fill_type="solid", start_color="92D050", end_color="92D050",)

# ======================================================
# LOGGING
# ======================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(levelname)s: %(message)s",
)

logger = logging.getLogger(__name__)

# ======================================================
# DATOVÉ TYPY
# ======================================================


@dataclass(slots=True)
class CategoryResult:
    category: str
    pdnyv_count: int
    third_sheet_count: int


# ======================================================
# EXCEL HELPERS
# ======================================================


def color_row(row: tuple[Cell, ...]) -> None:
    
    """
    Obarví všechny buňky v zadaném řádku.

    Nastaví žlutou barvu písma a černé pozadí.
    Používá se pro zvýraznění řádků mimo vybrané období.
    """
    for cell in row:
        cell.font = YELLOW_FONT
        cell.fill = BLACK_FILL


def get_month_from_value(value: object) -> str | None:
    
    """
    Získá měsíc ze zadané hodnoty.

    Podporuje objekty datetime i text ve formátu
    DD.MM.RRRR. Při neplatné hodnotě vrací None.
    """


    if value is None:
        return None

    if isinstance(value, datetime):
        return f"{value.month:02d}"

    try:
        parsed = datetime.strptime(
            str(value).strip(),
            "%d.%m.%Y",
        )
    except ValueError:
        return None

    return f"{parsed.month:02d}"

# ======================================================
# WORKBOOK
# ======================================================


def process_workbook(file_path: Path,) -> tuple[int, int]:
    
    """
    Zpracuje jeden Excel workbook.

    Obarví řádky neodpovídající vybraným měsícům,
    spočítá požadované hodnoty a případně uloží změny.

    Returns:
        Dvojici počtů pro list POUZE_pdnyv
        a třetí datový list.
    """
    try:
        wb = load_workbook(file_path, keep_vba=file_path.suffix.lower() == ".xlsm",)

    except Exception as exc:

        logger.error("Nelze otevřít %s: %s", file_path, exc,)

        return 0, 0
    
    try:

        workbook_modified = False

        if len(wb.worksheets) >= 2:

            for ws in wb.worksheets[FIRST_DATA_SHEET:LAST_DATA_SHEET]:

                for row in ws.iter_rows(min_row=DATA_START_ROW):

                    month = get_month_from_value( row[MONTH_COLUMN_INDEX].value)

                    if month is None:
                        continue

                    if month not in SELECTED_MONTHS:

                        color_row(row)

                        workbook_modified = True
        else:

            logger.warning( "Soubor %s nemá dostatek listů.", file_path.name,)

        pdnyv_count = 0
        third_sheet_count = 0

        if "POUZE_pdnyv" in wb.sheetnames:

            pdnyv_count = count_valid_rows(wb["POUZE_pdnyv"])

        if len(wb.worksheets) >= 3:

            third_sheet_count = count_valid_rows(wb.worksheets[THIRD_SHEET_INDEX])

        if workbook_modified:

            wb.save(file_path)

            logger.info(
                "Uložen: %s",
                file_path.name,
            )
    
    finally:
        wb.close()
    return (
        pdnyv_count,
        third_sheet_count,
    )
    
    
# ======================================================
# HELIKOPTÉROVÝ POHLED
# ======================================================
def create_helicopter_report(results: list[CategoryResult],) -> None:

    
    """
    Vytvoří souhrnný report za všechny kategorie.

    Report obsahuje počty z listu POUZE_pdnyv
    a z třetího listu každé zpracované kategorie.

    Výstup ukládá do samostatného Excel souboru.
    """
    HELICOPTER_FOLDER.mkdir(parents=True, exist_ok=True,)

    output_file = ( HELICOPTER_FOLDER / "helikopterovy_pohled.xlsx")

    wb = Workbook()
    ws = wb.active

    ws.title = "Helikoptérový pohled"

    ws["A1"] = "Kategorie"
    ws["B1"] = "POUZE_pdnyv"
    ws["C1"] = "3. list"

    for row_idx, result in enumerate(
        results,
        start=2,
    ):

        ws.cell(
            row=row_idx,
            column=1,
        ).value = result.category

        cell_pdnyv = ws.cell(
            row=row_idx,
            column=2,
        )

        if result.pdnyv_count == 0:
            cell_pdnyv.fill = GREEN_FILL
        else:
            cell_pdnyv.value = result.pdnyv_count

        cell_third = ws.cell(
            row=row_idx,
            column=3,
        )

        if result.third_sheet_count == 0:
            cell_third.fill = GREEN_FILL
        else:
            cell_third.value = (
                result.third_sheet_count
            )

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18

    wb.save(output_file)
    wb.close()

    logger.info(
        "Vytvořen přehled: %s",
        output_file,
    )
# ======================================================
# SOUBORY A SLOŽKY
# ======================================================


def find_latest_data_folder(category_folder: Path,) -> Path | None:
    
    """
    Vyhledá nejnovější DATA složku v kategorii.

    Z dostupných podsložek končících na '_DATA'
    vybere tu s nejnovějším časem změny.

    Returns:
        Cestu ke složce nebo None.
    """

    if not category_folder.is_dir():

        logger.warning( "Složka neexistuje: %s", category_folder,)

        return None

    data_folders = [
        folder
        for folder in category_folder.iterdir()
        if folder.is_dir()
        and folder.name.endswith("_DATA")
    ]

    if not data_folders:
        return None
    
    
    latest_folder = max( data_folders, key=lambda p: p.stat().st_mtime,)

    logger.info(
    "Kategorie %s -> nalezena nejnovější DATA složka: %s",
    category_folder.name,
    latest_folder,
)


    return latest_folder


def get_source_folder(category: str,) -> Path | None:

    """
    Vrátí zdrojovou složku pro zadanou kategorii.

    Nejprve sestaví cestu ke kategorii a následně
    vyhledá její nejnovější DATA složku.

    Returns:
        Zdrojovou složku nebo None.
    """
    category_folder = (SOURCE_ROOT / f"porovnavač_{category}" )

    return find_latest_data_folder(category_folder)


def get_target_category(category: str,) -> str:

    """
    Upraví název cílové kategorie podle pravidel.

    Kategorie začínající textem DROBY získají
    na začátek podtržítko.

    Returns:
        Název cílové kategorie.
    """
    if category.startswith("DROBY"):
        return f"_{category}"

    return category


def create_target_folder(category: str,) -> Path:

    """
    Vytvoří cílovou složku pro danou kategorii.

    Pokud složka neexistuje, bude automaticky
    vytvořena včetně nadřazených adresářů.

    Returns:
        Cestu k cílové složce.
    """
    target_folder = (TARGET_ROOT / get_target_category(category) )

    target_folder.mkdir( parents=True, exist_ok=True,)

    return target_folder


def find_category_files(source_folder: Path, category: str,) -> list[Path]:

    """
    Vyhledá soubory odpovídající dané kategorii.

    Hledání probíhá pomocí masky názvu souboru
    ve zdrojové složce kategorie.

    Returns:
        Seřazený seznam nalezených souborů.
    """   
    pattern = (f"vysledek_porovnani_{category}*")
    return sorted(source_folder.glob(pattern) )


def copy_file(source_file: Path,target_file: Path,) -> None:
    
    """
    Zkopíruje soubor do cílového umístění.

    Zachová metadata souboru pomocí funkce copy2
    a zapíše informaci do logu.

    Returns:
        None.
    """
    shutil.copy2( source_file, target_file,)

    logger.info(
        "%s -> %s",
        source_file.name,
        target_file.parent,
    )

def count_valid_rows(ws: Worksheet) -> int:

    """
    Spočítá platné řádky v pracovním listu.

    Do výsledku zahrnuje pouze řádky s vyplněnou
    první buňkou a vybraným měsícem.

    Returns:
        Počet vyhovujících řádků.
    """


    count = 0

    for row in ws.iter_rows(min_row=DATA_START_ROW):

        first_value = row[0].value

        if first_value in (None, ""):
            continue

        month = get_month_from_value(
            row[MONTH_COLUMN_INDEX].value
        )

        if month not in SELECTED_MONTHS:
            continue

        count += 1

    return count
# ======================================================
# ZPRACOVÁNÍ KATEGORIÍ
# ======================================================


def process_file(source_file: Path, target_folder: Path,) -> tuple[int, int]:
    
    """
    Zpracuje jeden zdrojový soubor.

    Soubor nejprve zkopíruje do cílové složky
    a v případě Excelu provede jeho analýzu.

    Returns:
        Dvojici zjištěných počtů.
    """


    target_file = (target_folder/ source_file.name)

    copy_file(source_file,target_file,)

    if target_file.suffix.lower() not in {
        ".xlsx",
        ".xlsm",
    }:
        return 0, 0

    return process_workbook(
        target_file
    )


def process_category_files(category: str, files: list[Path],) -> CategoryResult:

    """
    Zpracuje všechny soubory dané kategorie.

    Každý soubor zkopíruje do cílové složky,
    případně provede analýzu workbooku a
    agreguje získané počty.

    Returns:
        Souhrnný výsledek za celou kategorii.
    """


    target_folder = create_target_folder(category)    
    results = [ process_file( file, target_folder,) for file in files ]
    total_pdnyv = sum( pdnyv_count for pdnyv_count, _ in results)
    total_third = sum( third_count for _, third_count in results)    

    return CategoryResult(
        category=category,
        pdnyv_count=total_pdnyv,
        third_sheet_count=total_third,
    )


def process_category(category: str,) -> CategoryResult:
    
    """
    Zpracuje jednu kategorii dat.

    Vyhledá zdrojovou složku, najde příslušné
    soubory a vrátí jejich souhrnné výsledky.

    Returns:
        Souhrnné statistiky kategorie.
    """


    source_folder = get_source_folder(category)

    if source_folder is None:

        return CategoryResult(
            category=category,
            pdnyv_count=0,
            third_sheet_count=0,
        )

    files = find_category_files(
        source_folder,
        category,
    )

    if not files:

        logger.warning(
            "Nebyl nalezen soubor pro kategorii %s",
            category,
        )

        return CategoryResult(
            category=category,
            pdnyv_count=0,
            third_sheet_count=0,
        )

    return process_category_files(
        category,
        files,
    )


# ======================================================
# MAIN
# ======================================================
def main() -> None:
    
    """
    Spustí hlavní zpracování aplikace.

    Postupně zpracuje všechny definované kategorie,
    vytvoří souhrnný report a zapíše průběh do logu.

    Returns:
        None.
    """
    helicopter_results: list[CategoryResult] = []

    for category in CATEGORIES:

        result = process_category(category)

        helicopter_results.append(result)

    create_helicopter_report(helicopter_results)

    logger.info("Hotovo.")


if __name__ == "__main__":
    main()