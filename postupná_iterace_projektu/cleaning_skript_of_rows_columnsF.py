import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
import os

# ===== nastavení =====
input_file = "pdnyv.xlsx"
LIMIT_PRO_OSCIS=90000
KOD_PRACVMES_ARBITER=901099000
KOD_PRACVMES_FAU=905098000
output_dir = "vystupy"
ZADANY_SLOUPEC_1="den"


# ===== složky =====
base_output_dir = "vystupy"
os.makedirs(base_output_dir, exist_ok=True)

run_timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
output_dir = os.path.join(base_output_dir, run_timestamp)
os.makedirs(output_dir, exist_ok=True)

time_suffix = datetime.now().strftime("%H_%M")

# ===== načtení =====
df = pd.read_excel(input_file, engine="openpyxl")

# ===== sloupce =====
col_A = df.columns[0]
col_I = "pracvmes"
col_date = ZADANY_SLOUPEC_1
time_cols = ["priche", "odche", "prichv","odchv","prichsm","odchsm","odprac","odpracz","dfond"]

# ===== datum =====
df[col_date] = pd.to_datetime(df[col_date], errors="coerce")

# ===== bezpečné převody časů =====
def normalize_time(val):
    if pd.isna(val):
        return None

    if isinstance(val, (int, float)):
        return datetime(1899, 12, 30) + timedelta(days=val)

    if isinstance(val, timedelta):
        return datetime(1899, 12, 30) + val

    try:
        t = pd.to_datetime(val, errors="coerce")
        if pd.notnull(t):
            return datetime(1899, 12, 30, t.hour, t.minute, t.second)
    except:
        pass

    return None


for col in time_cols:
    if col in df.columns:
        df[col] = df[col].apply(normalize_time)

# ===== filtr =====
#df = df[(~df[col_I].isin([90, 80])) & (df[col_A] > 50)]
print("Filter")
df = df[(~df[col_I].isin([KOD_PRACVMES_ARBITER, KOD_PRACVMES_FAU])) & (df[col_A] < LIMIT_PRO_OSCIS)]


print("Prázdných příchodů")
if "prichv" in df.columns:
    df_prichv_empty = df[df["prichv"].isna()]
else:
    df_prichv_empty = pd.DataFrame()


# FUNKCE: filtr na platné časy
def has_valid_time(df_input):
    return df_input[
        (df_input["priche"].notna()) &
        (df_input["odche"].notna())
    ]

# ===== květen =====
df_may = df[df[col_date].dt.month == 5]

# ===== svátky =====
df_1_may = has_valid_time(df_may[df_may[col_date].dt.day == 1])
df_8_may = has_valid_time(df_may[df_may[col_date].dt.day == 8])

df_svatky = pd.concat([df_1_may, df_8_may], ignore_index=True)

# ===== víkendy =====
df_saturday = has_valid_time(df_may[df_may[col_date].dt.weekday == 5])
df_sunday = has_valid_time(df_may[df_may[col_date].dt.weekday == 6])

df_vikend= pd.concat([df_saturday, df_sunday])

# ===== export =====
def save_with_footer(df_input, base_filename):

    # pokud není co uložit, přeskočit
    if df_input.empty:
        print(f"Přeskočeno (žádná data): {base_filename}")
        return

    df_copy = df_input.copy()

    df_copy[col_date] = df_copy[col_date].dt.strftime("%d.%m.%Y")

    filename = f"{base_filename}_{time_suffix}.xlsx"
    filepath = os.path.join(output_dir, filename)

    df_copy.to_excel(filepath, index=False)

    wb = load_workbook(filepath)
    ws = wb.active

    # formát času
    for col_idx, col_name in enumerate(df_copy.columns, 1):
        if col_name in time_cols:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).number_format = "HH:MM"

    # footer
    footer_row = ws.max_row + 10
    now_full = datetime.now().strftime("%d.%m.%Y %H:%M:%S")

    ws.cell(row=footer_row, column=1, value=f"Vytvořeno: {now_full}")
    ws.cell(row=footer_row + 1, column=1,
            value="Softwarové závod sekce Státní tajemník MinFIN")

    wb.save(filepath)

# ===== exporty =====
save_with_footer(df_1_may, "vystup_1_kveten")
save_with_footer(df_8_may, "vystup_8_kveten")
save_with_footer(df_svatky, "vystup_svatky_kveten")

print("Export vikendů")
save_with_footer(df_saturday, "vystup_soboty_kveten")
save_with_footer(df_sunday, "vystup_nedele_kveten")
save_with_footer(df_vikend, "pdnyv_vikend.xlsx")

print("Export celeho výstupu")
save_with_footer(df, "pdnyv_vystup.xlsx")

print("Export prázdných příchodů")
save_with_footer(df_prichv_empty, "pdnyv_vystup_FILTER_pric")
                 

# ===== výpis =====
saturdays = sorted(df_saturday[col_date].dt.date.unique())
sundays = sorted(df_sunday[col_date].dt.date.unique())

print("Soboty v květnu:")
for d in saturdays:
    print(pd.to_datetime(d).strftime("%d.%m.%Y"))

print("\nNeděle v květnu:")
for d in sundays:
    print(pd.to_datetime(d).strftime("%d.%m.%Y"))

print("\nHotovo.")
print(f"Soubory jsou ve složce: {output_dir}")