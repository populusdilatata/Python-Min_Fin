import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
#import os
from pathlib import Path
from pandas.tseries.offsets import MonthEnd
from dataclasses import dataclass

# ===== nastavení =====
#input_file = "test_data2.xlsx"
input_file = "vstup_dochzarv2/06_DATA/dochzarv_06B.xlsx"


@dataclass(frozen=True)
class Config:
    limit_oscis: int = 90000
    arbiter: int = 901099000
    fau: int = 905098000

CFG = Config()

# ===== sloupce =====
@dataclass(frozen=True)
class Columns:
    start = "zapl"    
    end = "kopl"
    person_id = "oscis"
    work_place = "pracv"
    relation_to_org = "vztahorg"
  

COL = Columns()

"""
LIMIT_PRO_OSCIS=90000
KOD_PRACVMES_ARBITER=901099000
KOD_PRACVMES_FAU=905098000
"""
#base_output_dir = "vstup_dochzarv2"
base_output_dir = Path("vstup_dochzarv2")
zpracovavany_mesic="06_26"

#max_end = pd.Timestamp(start.year, 6, 30)
# ===== složky =====


output_dir = (base_output_dir /datetime.now().strftime("%Y-%m-%d_%H-%M") )


#os.makedirs(base_output_dir, exist_ok=True)

##run_timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
#output_dir = os.path.join(base_output_dir, run_timestamp)

output_dir.mkdir(parents=True, exist_ok=True)
#os.makedirs(output_dir, exist_ok=True)

time_suffix = datetime.now().strftime("%H_%M")




time_cols = ["zapl", "kopl"]
# ===== načtení =====
df = pd.read_excel(input_file, engine="openpyxl")

#col_A = df.columns[0]
#col_J = ZADANY_SLOUPEC_A

# ===== datum =====
#df[COL.start] = pd.to_datetime(df[COL.start], dayfirst=True, errors="coerce")
#df[COL.start] = pd.to_datetime(df[COL.start], format='%d.%m.%Y')



df[COL.start] = pd.to_datetime(df[COL.start], format="%d.%m.%Y", errors="coerce", )

#df[ZADANY_SLOUPEC_2] = pd.to_datetime(df[ZADANY_SLOUPEC_2], dayfirst=True,errors="coerce")


#df[ZADANY_SLOUPEC_1] = df[ZADANY_SLOUPEC_1].dt.strftime("%d.%m.%Y")
#df[ZADANY_SLOUPEC_2] = df[ZADANY_SLOUPEC_2].dt.strftime("%d.%m.%Y")

#df[ZADANY_SLOUPEC_1] = pd.to_datetime(df[ZADANY_SLOUPEC_1], format='%d.%m.%Y')

# ===== filtr =====

print("Filter")
MIN_START = pd.Timestamp(2026, 6, 1)
MAX_END = pd.Timestamp(2026, 6, 30)

#limit = pd.to_datetime('30.06.2026', format='%d.%m.%Y')
"""
df = df[(~df[COL.work_place].isin([KOD_PRACVMES_ARBITER, KOD_PRACVMES_FAU])) & (df[COL.person_id] < LIMIT_PRO_OSCIS)
        
      & (
        df[COL.start].isna()
        | (df[COL.start] <= MAX_END)
        )
]
"""

# Odfiltruj nevyhovující záznamy podle typu pracovního vztahu,
# hodnoty OSCIS 
# data zániku pracovního poměru.

valid_pracv = ~df[COL.work_place].isin([CFG.arbiter, CFG.fau])
valid_oscis = df[COL.person_id] < CFG.limit_oscis
valid_date = (df[COL.start].isna()| 
              (df[COL.start] <= MAX_END)
             )

df = df[valid_pracv & valid_oscis & valid_date]

def spojit_zaznamy(df, start_col, end_col, person_id_col):
    df = df.sort_values([person_id_col, start_col]).copy()

    ke_smazani = []

    for oscis, skupina in df.groupby(person_id_col):

        idxs = list(skupina.index)

        # Vytvoř dvojice po sobě jdoucích indexů (aktuální, následující)
        for idx1, idx2 in zip(idxs, idxs[1:]):
        #for i in range(len(idxs) - 1):

            #idx1 = idxs[i]
            #idx2 = idxs[i + 1]

            od1 = df.loc[idx1, start_col]
            do1 = df.loc[idx1, end_col]

            od2 = df.loc[idx2, start_col]
            do2 = df.loc[idx2, end_col]

            if pd.isna(do1) or pd.isna(od2):
                continue

            # navazující období
            if do1 + pd.Timedelta(days=1) == od2:

                print("\n--------------------------------")
                print(f"Osobní číslo: {oscis}")
                print(f"1. {od1.date()} - {do1.date()}")
                print(f"2. {od2.date()} - {do2.date()}")
                print("--------------------------------")

                volba = input("Spojit záznamy? (S/N): ").upper()

                if volba == "S":
                    df.loc[idx1, end_col] = do2
                    ke_smazani.append(idx2)

    df = df.drop(ke_smazani)

    return df


# Omez interval na sledované období:
# - datum začátku posuň nejdříve na MIN_START
# - chybějící datum konce nahraď MAX_END
# - datum konce posuň nejpozději na MAX_END

df[COL.start] = df[COL.start].clip(lower=MIN_START)

df[COL.end] = ( df[COL.end] 
    .fillna(MAX_END)
    .clip(upper=MAX_END)
)
#print(df)
df_spojene = spojit_zaznamy(df, COL.start, COL.end, COL.person_id)

duplicity = df_spojene[df_spojene.duplicated(subset=[COL.person_id], keep=False)]

print(duplicity)


# ===== filtr vztahorg = 3 =====
if COL.relation_to_org in df_spojene.columns:
    df_vztahorg_FILTER = df_spojene[df_spojene[COL.relation_to_org]==3]
    df_vztahorg_OSTATNI = df_spojene[df_spojene[COL.relation_to_org]!=3]






# ===== export =====
def save_with_footer(df_input, base_filename):

    if df_input.empty:
        print(f"Přeskočeno: {base_filename}")
        return

    df_copy = df_input.copy()
    df[COL.start] = pd.to_datetime(df_copy[COL.start], format='%d.%m.%Y')
    df_copy[COL.start] = df_copy[COL.start].dt.strftime("%d.%m.%Y")

    df[COL.end] = pd.to_datetime(df_copy[COL.end], format='%d.%m.%Y')
    df_copy[COL.end] = df_copy[COL.end].dt.strftime("%d.%m.%Y")

    filename = f"{base_filename}_{time_suffix}.xlsx"
    #filepath = os.path.join(output_dir, filename)
    filepath = output_dir / filename

    #df_input.to_excel(filepath, index=False)
    df_copy.to_excel(filepath, index=False)
    wb = load_workbook(filepath)
    ws = wb.active

    # formát času
    for col_idx, col_name in enumerate(df_input.columns, 1):
        if col_name in time_cols:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).number_format = "HH:MM"
    """
    # footer
    footer_row = ws.max_row + 10
    now_full = datetime.now().strftime("%d.%m.%Y %H:%M:%S")

    ws.cell(row=footer_row, column=1, value=f"Vytvořeno: {now_full}")
    ws.cell(row=footer_row + 1, column=1,
            value="Softwarové závod sekce Státní tajemník MinFIN")

    wb.save(filepath)
    """

# ===== exporty =====
# nový výstup dle FILTER + COUNTIF
print("Export filtrovaných záznamyů")
#print("-vztahorg = 3")
#print(df_vztahorg_FILTER)
#print("ostatních")
#print(df_vztahorg_OSTATNI)


save_with_footer(
    duplicity,
    "dochzarv2_"+zpracovavany_mesic+"_duplicity"
)

save_with_footer(
    df_vztahorg_FILTER,
    "dochzarv2_"+zpracovavany_mesic+"_FILTER_vztahorg"
)

save_with_footer(
    df_vztahorg_OSTATNI,
    "dochzarv2_"+zpracovavany_mesic+"_FILTER_vztahorg_NOT3"
)




# ===== výpis =====
print(f"\nHotovo.")
print(f"Soubory jsou ve složce: {output_dir}")