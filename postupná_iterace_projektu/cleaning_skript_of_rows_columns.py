import pandas as pd
from datetime import datetime
from openpyxl import load_workbook

# ===== nastavení =====
input_file = "pdnyv.xlsx"
LIMIT_PRO_OSCIS=90000
KOD_PRACVMES_ARBITER=901099000
KOD_PRACVMES_FAU=905098000


# ===== načtení =====
df = pd.read_excel(input_file, engine="openpyxl")

# sloupce
col_A = "oscis"
col_CP = "pracvmes"
col_date = "den"

# převod na datetime
df[col_date] = pd.to_datetime(df[col_date], errors="coerce")

# pokus o převod časových sloupců (pokud existují)
time_cols = ["priche", "odche"]
for col in time_cols:
    if col in df.columns:
        df[col] = pd.to_datetime(df[col], errors="coerce")


# ===== základní filtr =====
print("Základní filtr")
#df = df[(df[col_CP] != 90) & (df[col_A] < 90)]
df = df[(~df[col_CP].isin([KOD_PRACVMES_ARBITER, KOD_PRACVMES_FAU])) & (df[col_A] < LIMIT_PRO_OSCIS)]

# ===== pouze květen =====
df_may = df[df[col_date].dt.month == 5]

# ===== svátky =====
df_1_may = df_may[df_may[col_date].dt.day == 1]
df_8_may = df_may[df_may[col_date].dt.day == 8]
df_svatky= pd.concat([df_1_may, df_8_may])

# ===== víkendy =====
df_saturday = df_may[df_may[col_date].dt.weekday == 5]
df_sunday = df_may[df_may[col_date].dt.weekday == 6]
df_vikend= pd.concat([df_saturday, df_sunday])


# ===== funkce pro uložení + footer =====
def save_with_footer(df_input, filename):
    df_copy = df_input.copy()

    # formát data
    df_copy[col_date] = df_copy[col_date].dt.strftime("%d.%m.%Y")

    # uložit do Excelu
    df_copy.to_excel(filename, index=False)

    # otevřít soubor a přidat footer
    wb = load_workbook(filename)
    ws = wb.active

    last_row = ws.max_row
    footer_start = last_row + 10

    now = datetime.now().strftime("%d.%m.%Y %H:%M:%S")

    ws.cell(row=footer_start, column=1, value=f"Vytvořeno: {now}")
    ws.cell(row=footer_start + 1, column=1, value="Softwarový závod 77/Budova B, sekce Státní tajemník MinFIN")

    wb.save(filename)

# ===== export =====

print("EXPORT")
print("EXPORT svatků")
save_with_footer(df_1_may, "pdnyv_1_kveten.xlsx")
save_with_footer(df_8_may, "pdnyv_8_kveten.xlsx")
save_with_footer(df_svatky, "pdnyv_svatky.xlsx")

print("EXPORT víkendových dní")
save_with_footer(df_saturday, "pdnyv_soboty_kveten.xlsx")
save_with_footer(df_sunday, "pdnyv_nedele_kveten.xlsx")
save_with_footer(df_vikend, "pdnyv_vikend.xlsx")

save_with_footer(df, "pdnyv_test_data2.xlsx")

# ===== výpis víkendů =====
saturdays = sorted(df_saturday[col_date].dt.date.unique())
sundays = sorted(df_sunday[col_date].dt.date.unique())

print("Soboty v květnu:")
for d in saturdays:
    print(pd.to_datetime(d).strftime("%d.%m.%Y"))

print("\nNeděle v květnu:")
for d in sundays:
    print(pd.to_datetime(d).strftime("%d.%m.%Y"))

print("\nHotovo.")