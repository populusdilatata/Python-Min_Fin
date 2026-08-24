
from pathlib import Path
import logging
import shutil
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font


# TIP 23 - všechny konstanty na jednom místě
CATEGORIES = [
    "ABSEN",
    "INDIV",
    "LEKAR",
    "DOVOL",
    "NEMOC",
    "DROBY_OCR",
    "DROBY_OTECD",
    "DROBY_PLACV",
    "DROBY_STUDV",
    "DROBY_SVZOZ",
]

SOURCE_ROOT = Path("porovnavač")
TARGET_ROOT = Path("PRO_Sko")

SELECTED_MONTHS = {"06"}

MONTH_COLUMN = 1

YELLOW_FONT = Font(color="FFFF00")
BLACK_FILL = PatternFill(
    fill_type="solid",
    start_color="000000",
    end_color="000000",
)

# TIP 15 - použití logging místo print
logging.basicConfig(
    level=logging.INFO,
    format="%(levelname)s: %(message)s"
)


# TIP 19 - pomocná funkce
def color_row(row):
    """Obarví celý řádek."""
    for cell in row:
        cell.font = YELLOW_FONT
        cell.fill = BLACK_FILL


def process_excel(file_path: Path):
    """
    Zpracuje Excel a obarví řádky,
    jejichž měsíc není v SELECTED_MONTHS.
    """

    try:  # TIP 18 - ošetření výjimek

        # TIP 16 - zachování maker u XLSM souborů
        wb = load_workbook(
            file_path,
            keep_vba=file_path.suffix.lower() == ".xlsm"
        )

    except Exception as exc:
        logging.error("Nelze otevřít %s: %s", file_path, exc)
        return

    modified = False  # TIP 17

    # TIP 21 - kontrola počtu listů
    if len(wb.worksheets) < 2:
        logging.warning(
            "Soubor %s nemá dostatek listů.",
            file_path.name
        )
        return

    # TIP 20 - práce s výřezem listů
    for ws in wb.worksheets[1:3]:

        for row in ws.iter_rows(min_row=2):

            month_value = row[MONTH_COLUMN].value

            if month_value is None:
                continue

            try:
                # TIP 13 - použití datetime namísto split(".")
                parsed_date = datetime.strptime(
                    str(month_value).strip(),
                    "%d.%m.%Y"
                )

                month = f"{parsed_date.month:02d}"

            except ValueError:
                logging.warning(
                    "Neplatné datum '%s' v listu %s",
                    month_value,
                    ws.title
                )
                continue

            if month not in SELECTED_MONTHS:
                color_row(row)
                modified = True

    # TIP 17 - ukládej pouze pokud došlo ke změně
    if modified:
        wb.save(file_path)
        logging.info("Uložen: %s", file_path.name)


def find_latest_data_folder(category_folder: Path):

    # TIP 4 - kontrola existence složky
    if not category_folder.is_dir():
        logging.warning(
            "Složka neexistuje: %s",
            category_folder
        )
        return None

    data_folders = sorted(
        d for d in category_folder.iterdir()
        if d.is_dir() and d.name.endswith("_DATA")
    )

    if not data_folders:
        return None

    # TIP 5 - vhodnější je max(), zde ponecháno pro názornost
    return max(data_folders)


# TIP 25 - hlavní vstupní bod programu
def main():

    for category in CATEGORIES:

        category_folder = (
            SOURCE_ROOT
            / f"porovnavač_{category}"
        )

        source_folder = find_latest_data_folder(
            category_folder
        )

        if source_folder is None:
            logging.warning(
                "Nebyla nalezena DATA složka pro %s",
                category
            )
            continue

        target_category = (
            f"_{category}"
            if category.startswith("DROBY")
            else category
        )

        target_folder = TARGET_ROOT / target_category

        target_folder.mkdir(
            parents=True,
            exist_ok=True
        )

        pattern = (
            f"vysledek_porovnani_{category}*"
        )

        files = list(source_folder.glob(pattern))

        if not files:
            logging.warning(
                "Soubor %s nebyl nalezen.",
                pattern
            )
            continue

        for file in files:

            target_file = target_folder / file.name

            shutil.copy2(file, target_file)

            logging.info(
                "%s -> %s",
                file.name,
                target_folder
            )

            if target_file.suffix.lower() in (
                ".xlsx",
                ".xlsm",
            ):
                process_excel(target_file)


if __name__ == "__main__":
    main()
