import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
import os

# ===== nastavení =====
#input_file = "test_data2.xlsx"
input_file = "pdnyv.xlsx"
LIMIT_PRO_OSCIS=90000
KOD_PRACVMES_ARBITER=901099000
KOD_PRACVMES_FAU=905098000
output_dir = "vystupy"
ZADANY_SLOUPEC_1="den"

# ===== složky =====
base_output_dir = "vystupy"
os.makedirs(base_output_dir, exist_ok=True)

run_timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
output_dir = os.path.join(base_output_dir, run_timestamp)
os.makedirs(output_dir, exist_ok=True)

time_suffix = datetime.now().strftime("%H_%M")

# ===== načtení =====
df = pd.read_excel(input_file, engine="openpyxl")

# ===== sloupce =====
col_A = df.columns[0]
col_I = "pracvmes"
col_date = ZADANY_SLOUPEC_1

# ===== datum =====
df[col_date] = pd.to_datetime(df[col_date], errors="coerce")

# ===== filtr =====
#print("Filter")
#df = df[(~df[col_I].isin([KOD_PRACVMES_ARBITER, KOD_PRACVMES_FAU])) & (df[col_A] < LIMIT_PRO_OSCIS)]

print("Filter")
df = df[(~df[col_I].isin([KOD_PRACVMES_ARBITER, KOD_PRACVMES_FAU])) & (df[col_A] < LIMIT_PRO_OSCIS)]



# ===== květen =====
df_may = df[df[col_date].dt.month == 5]

# ===== svátky =====
df_1_may = df_may[df_may[col_date].dt.day == 1]
df_8_may = df_may[df_may[col_date].dt.day == 8]
df_svatky = pd.concat([df_1_may, df_8_may], ignore_index=True)

# ===== víkendy =====
df_saturday = df_may[df_may[col_date].dt.weekday == 5]
df_sunday = df_may[df_may[col_date].dt.weekday == 6]
df_vikend= pd.concat([df_saturday, df_sunday])



# ===== UNIVERZÁLNÍ převod času (klíčová část) =====
def format_excel_time(val):
    if pd.isna(val):
        return ""

    # číslo z Excelu (zlomek dne)
    if isinstance(val, (int, float)):
        seconds = int(val * 86400)
        h = seconds // 3600
        m = (seconds % 3600) // 60
        return f"{h:02}:{m:02}"

    # text nebo datetime
    try:
        t = pd.to_datetime(val, errors="coerce")
        if pd.notnull(t):
            return t.strftime("%H:%M")
    except:
        pass

    return ""


# ===== export =====
def save_with_footer(df_input, base_filename):
    df_copy = df_input.copy()

    # datum
    df_copy[col_date] = df_copy[col_date].dt.strftime("%d.%m.%Y")

    # čas – převod AŽ TADY
    for col in ["priche", "odche"]:
        if col in df_copy.columns:
            df_copy[col] = df_copy[col].apply(format_excel_time)

    # cesta
    filename = f"{base_filename}_{time_suffix}.xlsx"
    filepath = os.path.join(output_dir, filename)

    # uložit
    df_copy.to_excel(filepath, index=False)

    # ===== footer =====
    wb = load_workbook(filepath)
    ws = wb.active

    footer_row = ws.max_row + 10
    now_full = datetime.now().strftime("%d.%m.%Y %H:%M:%S")

    ws.cell(row=footer_row, column=1, value=f"Vytvořeno: {now_full}")
    ws.cell(row=footer_row + 1, column=1,
            value="Softwarové závod sekce Státní tajemník MinFIN")

    wb.save(filepath)


# ===== exporty =====
print("Export")
save_with_footer(df_1_may, "vystup_1_kveten")
save_with_footer(df_8_may, "vystup_8_kveten")
save_with_footer(df_svatky, "vystup_svatky_kveten")

print("Export vikendů")
save_with_footer(df_saturday, "vystup_soboty_kveten")
save_with_footer(df_sunday, "vystup_nedele_kveten")
save_with_footer(df_vikend, "pdnyv_vikend.xlsx")

print("Export celeho výstupu")
save_with_footer(df, "pdnyv_vystup.xlsx")

# ===== výpis =====
print("Výpis")
saturdays = sorted(df_saturday[col_date].dt.date.unique())
sundays = sorted(df_sunday[col_date].dt.date.unique())

print("Soboty v květnu:")
for d in saturdays:
    print(pd.to_datetime(d).strftime("%d.%m.%Y"))

print("\nNeděle v květnu:")
for d in sundays:
    print(pd.to_datetime(d).strftime("%d.%m.%Y"))

print(f"\nHotovo.")
print(f"Soubory: {output_dir}")
