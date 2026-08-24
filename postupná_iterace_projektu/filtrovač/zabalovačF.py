from pathlib import Path
import logging
import shutil
from datetime import datetime

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font

from dataclasses import dataclass


@dataclass(slots=True)
class CategoryResult:
    category: str
    pdnyv_count: int
    third_sheet_count: int

# ======================================================
# KONSTANTY
# ======================================================

FIRST_DATA_SHEET = 1
LAST_DATA_SHEET = 3


CATEGORIES = [
    "ABSEN",
    "INDIV",
    "LEKAR",
    "DOVOL",
    "NEMOC",
    "DROBY_OCR",
    "DROBY_OTECD",
    "DROBY_PLACV",
    "DROBY_STUDV",
    "DROBY_SVZOZ",
]

SOURCE_ROOT = Path("porovnavač")
TARGET_ROOT = Path("PRO_Sko")

HELICOPTER_FOLDER = Path("helikoptérový pohled")

SELECTED_MONTHS = {"06"}

# druhý sloupec = index 1
MONTH_COLUMN = 1

YELLOW_FONT = Font(color="FFFF00")

"""
BLACK_FILL = PatternFill(
    fill_type="solid",
    start_color="000000",
    end_color="000000",
)

GREEN_FILL = PatternFill(
    fill_type="solid",
    start_color="92D050",
    end_color="92D050",
)
"""

def make_fill(color: str) -> PatternFill:
    return PatternFill(
        fill_type="solid",
        start_color=color,
        end_color=color,
    )

BLACK_FILL = make_fill("000000")
GREEN_FILL = make_fill("92D050")


BLACK_RGB_VALUES = {
    "000000",
    "00000000",
    "FF000000",
}

YELLOW_RGB_VALUES = {
    "FFFF00",
    "00FFFF00",
    "FFFFFF00",
}


logging.basicConfig(
    level=logging.INFO,
    format="%(levelname)s: %(message)s"
)





# ======================================================
# BARVENÍ ŘÁDKŮ
# ======================================================

def color_row(row):
    """Obarví celý řádek černo-žlutě."""

    for cell in row:
        cell.font = YELLOW_FONT
        cell.fill = BLACK_FILL


def get_month_from_value(value) -> str | None:
    """
    Získá měsíc z textového nebo datového pole.
    """

    if value is None:
        return None

    if isinstance(value, datetime):
        return f"{value.month:02d}"
    
    """
    parsed = datetime.strptime(
        str(value).strip(),
        "%d.%m.%Y"
    )

    return f"{parsed.month:02d}"
    """
    
    try:
        parsed = datetime.strptime(
        str(value).strip(),
        "%d.%m.%Y"
    )
    except ValueError:
        return None

    return f"{parsed.month:02d}"


def process_workbook(file_path: Path) -> tuple[int, int]:

    wb = load_workbook(
        file_path,
        keep_vba=file_path.suffix.lower() == ".xlsm"
    )

    workbook_modified = False

    pdnyv_count = 0
    third_sheet_count = 0

    if len(wb.worksheets) >= 2:

        for sheet_index, ws in enumerate(
            wb.worksheets[1:3],
            start=1
        ):

            for row in ws.iter_rows(min_row=2):

                month_value = row[MONTH_COLUMN].value

                if month_value is None:
                    continue

                try:
                    month = get_month_from_value(
                        month_value
                    )
                except ValueError:
                    continue

                row_is_excluded = (
                    month not in SELECTED_MONTHS
                )

                if row_is_excluded:
                    color_row(row)
                    workbook_modified = True
                    continue

                # ==================================
                # POČÍTÁNÍ ŘÁDKŮ PŘÍMO TADY
                # ==================================

                first_value = row[0].value

                if first_value not in (None, ""):

                    if (
                        ws.title == "POUZE_pdnyv"
                    ):
                        pdnyv_count += 1

                    elif (
                        len(wb.worksheets) >= 3
                        and ws == wb.worksheets[2]
                    ):
                        third_sheet_count += 1

    if workbook_modified:
        wb.save(file_path)
        wb.close()

    return (
        pdnyv_count,
        third_sheet_count
    )

def create_helicopter_report(results):

    HELICOPTER_FOLDER.mkdir(
        parents=True,
        exist_ok=True
    )

    output_file = (
        HELICOPTER_FOLDER
        / "helikopterovy_pohled.xlsx"
    )

    wb = Workbook()
    ws = wb.active

    ws.title = "Helikoptérový pohled"

    ws["A1"] = "Kategorie"
    ws["B1"] = "POUZE_pdnyv"
    ws["C1"] = "3. list"

    for row_idx, (
        category,
        pdnyv_count,
        third_sheet_count
    ) in enumerate(results, start=2):

        ws.cell(
            row=row_idx,
            column=1
        ).value = category

        cell_pdnyv = ws.cell(
            row=row_idx,
            column=2
        )

        if pdnyv_count == 0:
            cell_pdnyv.fill = GREEN_FILL
        else:
            cell_pdnyv.value = pdnyv_count

        cell_third = ws.cell(
            row=row_idx,
            column=3
        )

        if third_sheet_count == 0:
            cell_third.fill = GREEN_FILL
        else:
            cell_third.value = third_sheet_count

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18

    wb.save(output_file)
    wb.close()

    logging.info(
        "Vytvořen přehled: %s",
        output_file
    )

# ======================================================
# DATA SLOŽKY
# ======================================================

def find_latest_data_folder(category_folder: Path):

    if not category_folder.is_dir():

        logging.warning(
            "Složka neexistuje: %s",
            category_folder
        )

        return None

    data_folders = [
        d
        for d in category_folder.iterdir()
        if d.is_dir()
        and d.name.endswith("_DATA")
    ]

    if not data_folders:
        return None

    return max(data_folders)


def find_category_files(source_folder: Path,category: str) -> listpattern = (
        f"vysledek_porovnani_{category}*"
    )

    files = sorted(
        source_folder.glob(pattern)
    )

    return files

def get_source_folder(category: str) -> Path | None:

    category_folder = (
        SOURCE_ROOT
        / f"porovnavač_{category}"
    )

    return find_latest_data_folder(
        category_folder
    )


def get_target_category(
    category: str
) -> str:

    if category.startswith("DROBY"):
        return f"_{category}"

    return category


def process_category(category: str) -> CategoryResult:

    source_folder = get_source_folder(
        category
    )

    if source_folder is None:

        return CategoryResult(
            category=category,
            pdnyv_count=0,
            third_sheet_count=0,
        )

    files = find_category_files(
        source_folder,
        category
    )

    if not files:

        logger.warning(
            "Soubor pro kategorii %s nebyl nalezen.",
            category
        )

        return CategoryResult(
            category=category,
            pdnyv_count=0,
            third_sheet_count=0,
        )

    return process_category_files(
        category,
        files
    )
# ======================================================
# MAIN
# ======================================================

def main():

    helicopter_results: list[CategoryResult] = []

    for category in CATEGORIES:

        result = process_category(category)

        helicopter_results.append(result)

    create_helicopter_report(
        helicopter_results
    )

    logger.info("Hotovo.")



if __name__ == "__main__":
    main()