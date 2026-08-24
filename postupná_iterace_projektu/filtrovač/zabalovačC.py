from pathlib import Path
import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

CATEGORIES = [
    "ABSEN",
    "INDIV",
    "LEKAR",
    "DOVOL",
    "NEMOC",

    "DROBY_OCR",
    "DROBY_OTECD",
    "DROBY_PLACV",
    "DROBY_STUDV",
    "DROBY_SVZOZ"
]

SOURCE_ROOT = Path("porovnavač")
TARGET_ROOT = Path("PRO_Sko")

# Příklad povolených měsíců
SELECTED_MONTHS = {"06"}  # leden–březen
MONTH_COLUMN = 1

# Styly
YELLOW_FONT = Font(color="FFFF00")      # žluté písmo
BLACK_FILL = PatternFill(
    fill_type="solid",
    start_color="000000",
    end_color="000000"
)

for category in CATEGORIES:

    category_folder = SOURCE_ROOT / f"porovnavač_{category}"

    
    if not category_folder.is_dir():
        print(f"Složka neexistuje nebo není adresář: {category_folder}")
        continue


    data_folders = [
        d for d in category_folder.iterdir()
         if d.is_dir() and d.name.endswith("_DATA")
                    ]
    

    if not data_folders:
        print(f"Nebyla nalezena DATA složka pro {category}.")
        continue

    #source_folder = data_folders[-1]
    source_folder = max(data_folders)

    target_category = f"_{category}" if category.startswith("DROBY") else category

    target_folder = TARGET_ROOT / target_category
    target_folder.mkdir(parents=True, exist_ok=True)

    pattern = f"vysledek_porovnani_{category}*"
    files = list(source_folder.glob(pattern))

    if len(files) == 0:
        print(f"Soubor {pattern} nebyl nalezen.")
        continue

    for file in files:
        target_file = target_folder / file.name

        shutil.copy2(file, target_file)
        print(f"{file.name} -> {target_folder}")

        # Úprava Excelu
        if target_file.suffix.lower() in (".xlsx", ".xlsm"):
            wb = load_workbook(target_file)

            for ws in wb.worksheets[1:3]:
                for row in ws.iter_rows(min_row=2):

                    month_value = row[MONTH_COLUMN].value  # 2. sloupec (B)
                    
                    if month_value is None:
                        continue

                    month_str = str(month_value).strip()

                    # očekává formát DD.MM.RRRR
                    parts = month_str.split(".")
                    print("Toto je měsíc")
                    print(parts[1])

                    #if month_value is not None and str(month_value) not in SELECTED_MONTHS:
                    if parts[1] is not None and str(parts[1]) not in SELECTED_MONTHS:       
                        # Obarvení CELÉHO řádku
                        for cell in row:
                            cell.font = YELLOW_FONT
                            cell.fill = BLACK_FILL

            wb.save(target_file)