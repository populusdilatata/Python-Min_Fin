from pathlib import Path
import logging
import shutil
from datetime import datetime

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font


# ======================================================
# KONSTANTY
# ======================================================

CATEGORIES = [
    "ABSEN",
    "INDIV",
    "LEKAR",
    "DOVOL",
    "NEMOC",
    "DROBY_OCR",
    "DROBY_OTECD",
    "DROBY_PLACV",
    "DROBY_STUDV",
    "DROBY_SVZOZ",
]

SOURCE_ROOT = Path("porovnavač")
TARGET_ROOT = Path("PRO_Sko")

HELICOPTER_FOLDER = Path("helikoptérový pohled")

SELECTED_MONTHS = {"06"}

MONTH_COLUMN = 1

YELLOW_FONT = Font(color="FFFF00")

BLACK_FILL = PatternFill(
    fill_type="solid",
    start_color="000000",
    end_color="000000",
)

GREEN_FILL = PatternFill(
    fill_type="solid",
    start_color="00FF00",
    end_color="00FF00",
)

logging.basicConfig(
    level=logging.INFO,
    format="%(levelname)s: %(message)s"
)


# ======================================================
# EXCEL BARVENÍ
# ======================================================

def color_row(row):
    """Obarví celý řádek."""
    for cell in row:
        cell.font = YELLOW_FONT
        cell.fill = BLACK_FILL


def process_excel(file_path: Path):
    """
    Zpracuje Excel a obarví řádky,
    jejichž měsíc není v SELECTED_MONTHS.
    """

    try:
        wb = load_workbook(
            file_path,
            keep_vba=file_path.suffix.lower() == ".xlsm"
        )

    except Exception as exc:
        logging.error(
            "Nelze otevřít %s: %s",
            file_path,
            exc
        )
        return

    modified = False

    if len(wb.worksheets) < 2:
        logging.warning(
            "Soubor %s nemá dostatek listů.",
            file_path.name
        )
        return

    for ws in wb.worksheets[1:3]:

        for row in ws.iter_rows(min_row=2):

            month_value = row[MONTH_COLUMN].value

            if month_value is None:
                continue

            try:
                parsed_date = datetime.strptime(
                    str(month_value).strip(),
                    "%d.%m.%Y"
                )

                month = f"{parsed_date.month:02d}"

            except ValueError:

                logging.warning(
                    "Neplatné datum '%s' v listu %s",
                    month_value,
                    ws.title
                )
                continue

            if month not in SELECTED_MONTHS:
                color_row(row)
                modified = True

    if modified:
        wb.save(file_path)
        logging.info(
            "Uložen: %s",
            file_path.name
        )


# ======================================================
# HELIKOPTÉROVÝ POHLED
# ======================================================

def count_non_empty_first_column(ws):
    """
    Spočítá neprázdné hodnoty v prvním sloupci
    od druhého řádku.
    """

    return sum(
        1
        for row in ws.iter_rows(
            min_row=2,
            max_col=1,
            values_only=True
        )
        if row[0] not in (None, "")
    )


def count_oscis(file_path: Path):
    """
    Vrací:
    (
        počet OSCIS v listu POUZE_pdnyv,
        počet OSCIS ve 3. listu
    )
    """

    try:
        wb = load_workbook(
            file_path,
            read_only=True,
            keep_vba=file_path.suffix.lower() == ".xlsm"
        )

    except Exception as exc:

        logging.error(
            "Nelze načíst %s: %s",
            file_path,
            exc
        )

        return 0, 0

    pdnyv_count = 0
    third_sheet_count = 0

    if "POUZE_pdnyv" in wb.sheetnames:

        ws = wb["POUZE_pdnyv"]

        pdnyv_count = count_non_empty_first_column(ws)

    if len(wb.worksheets) >= 3:

        ws = wb.worksheets[2]

        third_sheet_count = count_non_empty_first_column(ws)

    return pdnyv_count, third_sheet_count


def create_helicopter_report(results):
    """
    Vytvoří souhrnný report.

    results = [
        (category, pdnyv_count, third_sheet_count)
    ]
    """

    HELICOPTER_FOLDER.mkdir(
        parents=True,
        exist_ok=True
    )

    output_file = (
        HELICOPTER_FOLDER
        / "helikopterovy_pohled.xlsx"
    )

    wb = Workbook()
    ws = wb.active

    ws.title = "Helikoptérový pohled"

    ws["A1"] = "Kategorie"
    ws["B1"] = "POUZE_pdnyv"
    ws["C1"] = "3. list"

    for row_idx, (
        category,
        pdnyv_count,
        third_sheet_count
    ) in enumerate(results, start=2):

        ws.cell(
            row=row_idx,
            column=1
        ).value = category

        pdnyv_cell = ws.cell(
            row=row_idx,
            column=2
        )

        if pdnyv_count == 0:
            pdnyv_cell.fill = GREEN_FILL
        else:
            pdnyv_cell.value = pdnyv_count

        third_cell = ws.cell(
            row=row_idx,
            column=3
        )

        if third_sheet_count == 0:
            third_cell.fill = GREEN_FILL
        else:
            third_cell.value = third_sheet_count

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20

    wb.save(output_file)

    logging.info(
        "Vytvořen souhrnný přehled: %s",
        output_file
    )


# ======================================================
# HLEDÁNÍ ZDROJOVÝCH SLOŽEK
# ======================================================

def find_latest_data_folder(category_folder: Path):

    if not category_folder.is_dir():

        logging.warning(
            "Složka neexistuje: %s",
            category_folder
        )

        return None

    data_folders = [
        folder
        for folder in category_folder.iterdir()
        if folder.is_dir()
        and folder.name.endswith("_DATA")
    ]

    if not data_folders:
        return None

    return max(data_folders)


# ======================================================
# MAIN
# ======================================================

def main():

    helicopter_results = []

    for category in CATEGORIES:

        category_folder = (
            SOURCE_ROOT
            / f"porovnavač_{category}"
        )

        source_folder = find_latest_data_folder(
            category_folder
        )

        if source_folder is None:

            logging.warning(
                "Nebyla nalezena DATA složka pro %s",
                category
            )

            helicopter_results.append(
                (category, 0, 0)
            )

            continue

        target_category = (
            f"_{category}"
            if category.startswith("DROBY")
            else category
        )

        target_folder = (
            TARGET_ROOT
            / target_category
        )

        target_folder.mkdir(
            parents=True,
            exist_ok=True
        )

        pattern = (
            f"vysledek_porovnani_{category}*"
        )

        files = list(
            source_folder.glob(pattern)
        )

        if not files:

            logging.warning(
                "Soubor %s nebyl nalezen.",
                pattern
            )

            helicopter_results.append(
                (category, 0, 0)
            )

            continue

        total_pdnyv = 0
        total_third = 0

        for file in files:

            target_file = (
                target_folder
                / file.name
            )

            shutil.copy2(
                file,
                target_file
            )

            logging.info(
                "%s -> %s",
                file.name,
                target_folder
            )

            pdnyv_count, third_count = (
                count_oscis(target_file)
            )

            total_pdnyv += pdnyv_count
            total_third += third_count

            if target_file.suffix.lower() in (
                ".xlsx",
                ".xlsm",
            ):
                process_excel(target_file)

        helicopter_results.append(
            (
                category,
                total_pdnyv,
                total_third
            )
        )

    create_helicopter_report(
        helicopter_results
    )


if __name__ == "__main__":
    main()