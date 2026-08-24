import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from pathlib import Path
from dataclasses import dataclass

# ===== nastavení =====
#input_file = "test_data2.xlsx"
input_file = "vstup_pdnyv/pdnyv_06.xlsx"

@dataclass(frozen=True)
class Config:
    limit_oscis: int = 90000
    arbiter: int = 901099000
    fau: int = 905098000
    #===========================
    saturday = 5
    sunday = 6
    min_empty_priche_count = 25
    #===========================
    year = 2026
    month = 6
CFG = Config()

base_output_dir = "porovnavač"
zpracovavany_mesic="06_26"

EXCEL_EPOCH = datetime(1899, 12, 30)
# ===== sloupce =====
ZADANY_SLOUPEC_1="den"
ZADANY_SLOUPEC_2="oscis"
ZADANY_SLOUPEC_3="pracvmes"
FILTER_ZAZNAMU_A="duvodt"

#time_cols = ["priche", "odche", "prichv","odchv","prichsm","odchsm","odprac","odpracz","dfond"]

TIME_COLS = {
    "priche",
    "odche",
    "prichv",
    "odchv",
    "prichsm",
    "odchsm",
    "odprac",
    "odpracz",
    "dfond"
}
#data_pro_rozdělení

CATEGORY_MAP= {
    "dovol": "dovol",
    "puldo": "dovol",

    "odpra": "odpra",
    "služ": "odpra",

    "absen": "absen",
    "nemoc": "nemoc",
    "indiv": "indiv",
    "lékař": "lékař"
}



# ===== složky =====
def create_output_directory():

    now = datetime.now()

    run_timestamp = now.strftime("%Y-%m-%d_%H-%M")
    time_suffix = now.strftime("%H_%M")

    base_output_dir = Path("porovnavač")
    base_output_dir.mkdir(exist_ok=True)

    output_dir = base_output_dir / run_timestamp
    output_dir.mkdir(exist_ok=True)

    return output_dir, time_suffix


# ===== načtení =====
def load_data():

    df = pd.read_excel( input_file, engine="openpyxl")

    df[ZADANY_SLOUPEC_1] = pd.to_datetime( df[ZADANY_SLOUPEC_1], errors="coerce")

    return df

# ===== bezpečný převod času =====
"""

def normalize_time(val):
    if pd.isna(val):
        return None

    if isinstance(val, (int, float)):
        return EXCEL_EPOCH + timedelta(days=val)

    if isinstance(val, timedelta):
        return EXCEL_EPOCH + val

    try:
        t = pd.to_datetime(val, errors="coerce")
        if pd.notnull(t):
            #return datetime(EXCEL_EPOCH.year, EXCEL_EPOCH.month, EXCEL_EPOCH.day, t.hour, t.minute, t.second)
            return EXCEL_EPOCH.replace(hour=t.hour, minute=t.minute, second=t.second)
    except Exception:
        pass

    return None

def normalize_time_columns(df):

    for col in TIME_COLS:
        if col in df.columns:
            df[col] = df[col].apply(normalize_time)

    return df


def normalize_time_columns(df):

    for col in TIME_COLS:

        if col not in df.columns:
            continue

        
        print(col)
        print(df[col].dtype)


        print(f"\n=== {col} ===")

        print(
            df[col]
            .map(type)
            .value_counts()
            .head(10)
        )

        df[col] = df[col].apply(normalize_time)

    return df
"""
def normalize_time_columns(df):

    for col in TIME_COLS:

        if col not in df.columns:
            continue

        
        if pd.api.types.is_timedelta64_dtype(df[col]):
            df[col] = EXCEL_EPOCH + df[col]


    return df

# ===== filtr =====
def apply_main_filter(df):

    filter_mask = (~df[ZADANY_SLOUPEC_3].isin([
                                                CFG.arbiter,
                                                CFG.fau
                                                ])
                    ) & (
                        df[ZADANY_SLOUPEC_2] < CFG.limit_oscis
                        )

    return df.loc[filter_mask]

# ===== filtr prázdné priche =====
"""
if "priche" in df.columns:
    df_prichv_empty = df[df["priche"].isna()]
else:
    df_prichv_empty = pd.DataFrame()
"""

def get_empty_priche(df: pd.DataFrame) -> pd.DataFrame:

    if "priche" not in df.columns:
        return pd.DataFrame()

    return df.loc[df["priche"].isna()]


# FUNKCE: filtr na platné časy
def has_valid_time(df_input):

    #Funkce odstrani záznamy, které NEMAJÍ vyplněn sloupeček "priche", nebo "odche"
   
    return df_input.dropna(subset=["priche", "odche"])





# ===== APLIKACE EXCEL LOGIKY =====
"""
if not df_prichv_empty.empty:
    counts = df_prichv_empty[ZADANY_SLOUPEC_2].value_counts()
    valid_values = counts[counts > 25].index

    df_prichv_empty_filtered = df_prichv_empty[
        df_prichv_empty[ZADANY_SLOUPEC_2].isin(valid_values)
    ]
else:
    df_prichv_empty_filtered = pd.DataFrame()
"""
def filter_frequent_empty_priche(df_prichv_empty: pd.DataFrame) -> pd.DataFrame:

    if df_prichv_empty.empty:
        return pd.DataFrame()

    counts = (
        df_prichv_empty[ZADANY_SLOUPEC_2]
        .value_counts()
    )

    valid_values = (
        counts[counts > CFG.min_empty_priche_count]
        .index
    )

    return df_prichv_empty.loc[
        df_prichv_empty[ZADANY_SLOUPEC_2]
        .isin(valid_values)
    ]

# definice svátků (měsíc: seznam dní)
def get_holidays(df):

    holiday_mask = (
        ((df[ZADANY_SLOUPEC_1].dt.month == 5) &
         (df[ZADANY_SLOUPEC_1].dt.day.isin([1, 8])))
        |
        ((df[ZADANY_SLOUPEC_1].dt.month == 7) &
         (df[ZADANY_SLOUPEC_1].dt.day.isin([5, 6])))
    )

    return df.loc[holiday_mask]


# definice sobot a neděl

def get_weekends(df):

    df_saturday = df.loc[df[ZADANY_SLOUPEC_1].dt.weekday == CFG.saturday]
    df_sunday = df.loc[df[ZADANY_SLOUPEC_1].dt.weekday == CFG.sunday]

    df_vikend = pd.concat([
        df_saturday,
        df_sunday
    ])

    return (
        df_saturday,
        df_sunday,
        df_vikend
    )
# ===== export =====
def build_filename(name):
    return f"pdnyv_vystup_{zpracovavany_mesic}_{name}"

def save_with_footer(df_input, base_filename):

    if df_input.empty:
        print(f"Přeskočeno: {base_filename}")
        return

    df_copy = df_input.copy()
    df_copy.loc[:, ZADANY_SLOUPEC_1] = ( df_copy[ZADANY_SLOUPEC_1].dt.strftime("%d.%m.%Y") )

    filename = f"{base_filename}_{time_suffix}.xlsx"
    #filepath = os.path.join(output_dir, filename)
    filepath = output_dir / filename

    df_copy.to_excel(filepath, index=False)

    wb = load_workbook(filepath)
    ws = wb.active

    # formát času
    for col_idx, col_name in enumerate(df_copy.columns, 1):
        if col_name in TIME_COLS:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).number_format = "HH:MM"
                
    wb.save(filepath)



def export_weekends(df_saturday,df_sunday):
    print("Export víkendů")
    df_saturday_valid = has_valid_time(df_saturday)

    save_with_footer(
        df_saturday_valid,
        build_filename("soboty")
    )

    df_sunday_valid = has_valid_time(
        df_sunday
    )

    save_with_footer(
        df_sunday_valid,
        build_filename("nedele")
    )

    df_vikend_valid = pd.concat(
        [
            df_saturday_valid,
            df_sunday_valid
        ]
    )

    save_with_footer(
        df_vikend_valid,
        build_filename("vikendy")
    )


def export_categories(df_clean):
    print("Exporty rozdělených do kategorií")
    for name, df_part in (df_clean.groupby("kategorie")):
        save_with_footer(
            df_part.drop(
                columns=["kategorie"]
            ),
            f"pdnyv_{zpracovavany_mesic}_FILTER_{name}"
        )
def export_holidays(df_svatky):
    print("Export svátků")
    save_with_footer(
        has_valid_time(df_svatky),
        build_filename("svatky")
    )


def export_empty_priche(df_prichv_empty_filtered):
    print("Export prázdných příchodů")
    save_with_footer(
        df_prichv_empty_filtered,
        build_filename("FILTER_prichv_EMPTY")
    )

def main():

    global output_dir
    global time_suffix

    output_dir, time_suffix = (create_output_directory())

    df = load_data()

    df = normalize_time_columns(df)

    df = apply_main_filter(df)
    
    df_prichv_empty = get_empty_priche(df)

    
    df_prichv_empty_filtered = (filter_frequent_empty_priche(df_prichv_empty))


    df_svatky = get_holidays(df)

    (
        df_saturday,
        df_sunday,
        df_vikend
    ) = get_weekends(df)

    print("Exporty")

    

    

    # ===== exporty =====
    # nový výstup dle FILTER + COUNTIF
      
    export_empty_priche(df_prichv_empty_filtered)

    """
    save_with_footer(
        df_prichv_empty_filtered,
        "pdnyv_vystup_"+zpracovavany_mesic+"_FILTER_prichv_EMPTY"
    )
    """
    excluded_idx = (
        set(df_vikend.index)
        | set(df_svatky.index)
        | set(df_prichv_empty_filtered.index)
    )

    df_clean = df.drop(index=excluded_idx)

    #df_clean = df.drop(df_excluded.index, errors="ignore")


    save_with_footer(
        df_clean,
        "pdnyv_porovnat_"+zpracovavany_mesic
    )

    #print("Export svátků")
    export_holidays(df_svatky)

    """
    df_svatky_valid =has_valid_time(df_svatky)

    save_with_footer(df_svatky_valid, build_filename("svatky"))
    """
    
    """
    df_saturday_valid=has_valid_time(df_saturday)
    save_with_footer(df_saturday_valid, build_filename("soboty"))
    df_sunday_valid=has_valid_time(df_sunday)
    save_with_footer(df_sunday_valid, build_filename("nedele"))
    df_vikend_valid = pd.concat([df_saturday_valid, df_sunday_valid], ignore_index=False)
    save_with_footer(df_vikend_valid, build_filename("vikendy"))
    """
    #print("Export vikendů")
    export_weekends(df_saturday, df_sunday)



    # ===== vytvoření cílového sloupce =====
    #print("Export rozdělených ")

    #df_clean["kategorie"] = df_clean[FILTER_ZAZNAMU_A].map(category_map)

    #df_clean = df.drop(df_excluded.index, errors="ignore").copy()
    df_clean = df.drop(index=excluded_idx).copy()

    df_clean.loc[:, "kategorie"] = (df_clean[FILTER_ZAZNAMU_A].map(CATEGORY_MAP).fillna("droby") )


    # neznámé hodnoty → droby
    #df_clean["kategorie"] = df_clean["kategorie"].fillna("droby")

    # ===== rozdělení =====
    
    """
    for name, df_part in df_clean.groupby("kategorie"):
        save_with_footer(
            df_part.drop(columns=["kategorie"]),
            f"pdnyv_{zpracovavany_mesic}_FILTER_{name}"
        )

    """
    export_categories(df_clean)
    # ===== výpis =====
    print(f"\nHotovo.")
    print(f"Soubory jsou ve složce: {output_dir}")


if __name__ == "__main__":
    main()
