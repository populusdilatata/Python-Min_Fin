import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
import os

# ===== nastavení =====
#input_file = "test_data2.xlsx"
input_file = "pdnyv.xlsx"
LIMIT_PRO_OSCIS=90000
KOD_PRACVMES_ARBITER=901099000
KOD_PRACVMES_FAU=905098000
output_dir = "vystupy"
ZADANY_SLOUPEC_1="den"


# ===== složky =====
base_output_dir = "vystupy"
os.makedirs(base_output_dir, exist_ok=True)

run_timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
output_dir = os.path.join(base_output_dir, run_timestamp)
os.makedirs(output_dir, exist_ok=True)

time_suffix = datetime.now().strftime("%H_%M")

# ===== načtení =====
df = pd.read_excel(input_file, engine="openpyxl")

# ===== sloupce =====
col_A = df.columns[0]
col_I = "pracvmes"
col_date = ZADANY_SLOUPEC_1
time_cols = ["priche", "odche", "prichv","odchv","prichsm","odchsm","odprac","odpracz","dfond"]

# ===== datum =====
df[col_date] = pd.to_datetime(df[col_date], errors="coerce")

# ===== bezpečný převod času =====
def normalize_time(val):
    if pd.isna(val):
        return None

    if isinstance(val, (int, float)):
        return datetime(1899, 12, 30) + timedelta(days=val)

    if isinstance(val, timedelta):
        return datetime(1899, 12, 30) + val

    try:
        t = pd.to_datetime(val, errors="coerce")
        if pd.notnull(t):
            return datetime(1899, 12, 30, t.hour, t.minute, t.second)
    except:
        pass

    return None

for col in time_cols:
    if col in df.columns:
        df[col] = df[col].apply(normalize_time)

# ===== filtr =====
#df = df[(~df[col_I].isin([90, 80])) & (df[col_A] > 50)]
#df = df[(~df[col_I].isin([90, 80])) & (df[col_A] > 50)]

print("Filter")
df = df[(~df[col_I].isin([KOD_PRACVMES_ARBITER, KOD_PRACVMES_FAU])) & (df[col_A] < LIMIT_PRO_OSCIS)]


# ===== filtr prázdné prichv =====
if "prichv" in df.columns:
    df_prichv_empty = df[df["prichv"].isna()]
else:
    df_prichv_empty = pd.DataFrame()

# FUNKCE: filtr na platné časy
def has_valid_time(df_input):
    #Funkce odstrani záznamy, které NEMAJÍ vyplněn sloupeček "priche", nebo "odche"
    #print("Filter na platné časy")
    return df_input[
        (df_input["priche"].notna()) &
        (df_input["odche"].notna())
    ]




# ===== ✅ APLIKACE EXCEL LOGIKY =====
if not df_prichv_empty.empty:
    counts = df_prichv_empty[col_A].value_counts()
    valid_values = counts[counts > 28].index

    df_prichv_empty_filtered = df_prichv_empty[
        df_prichv_empty[col_A].isin(valid_values)
    ]
else:
    df_prichv_empty_filtered = pd.DataFrame()

# ===== květen =====
df_may = df[df[col_date].dt.month == 5]

df_1_may = df_may[df_may[col_date].dt.day == 1]
df_8_may = df_may[df_may[col_date].dt.day == 8]
df_svatky = pd.concat([df_1_may, df_8_may], ignore_index=False)

df_saturday = df_may[df_may[col_date].dt.weekday == 5]
df_sunday = df_may[df_may[col_date].dt.weekday == 6]
df_vikend= pd.concat([df_saturday, df_sunday])

# ===== export =====
def save_with_footer(df_input, base_filename):

    if df_input.empty:
        print(f"Přeskočeno: {base_filename}")
        return

    df_copy = df_input.copy()
    df_copy[col_date] = df_copy[col_date].dt.strftime("%d.%m.%Y")

    filename = f"{base_filename}_{time_suffix}.xlsx"
    filepath = os.path.join(output_dir, filename)

    df_copy.to_excel(filepath, index=False)

    wb = load_workbook(filepath)
    ws = wb.active

    # formát času
    for col_idx, col_name in enumerate(df_copy.columns, 1):
        if col_name in time_cols:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).number_format = "HH:MM"

    # footer
    footer_row = ws.max_row + 10
    now_full = datetime.now().strftime("%d.%m.%Y %H:%M:%S")

    ws.cell(row=footer_row, column=1, value=f"Vytvořeno: {now_full}")
    ws.cell(row=footer_row + 1, column=1,
            value="Softwarové závod sekce Státní tajemník MinFIN")

    wb.save(filepath)

# ===== exporty =====
# nový výstup dle FILTER + COUNTIF
print("Export prázdných příchodů")

save_with_footer(
    df_prichv_empty_filtered,
    "pdnyv_vystup_FILTER_prichv_EMPTY"
)


df_excluded = pd.concat([
    df_vikend,
    df_svatky,
    df_prichv_empty_filtered
], ignore_index=False)

df_clean = df.drop(df_excluded.index, errors="ignore")
save_with_footer(
    df_clean,
    "pdnyv_meziclanek"
)

print("Export svátků")
df_1_may_valid =has_valid_time(df_1_may)
save_with_footer(df_1_may_valid, "vystup_1_kveten")
df_8_may_valid =has_valid_time(df_8_may)
save_with_footer(df_8_may_valid, "vystup_8_kveten")
df_svatky_valid = pd.concat([df_1_may_valid, df_8_may_valid], ignore_index=False)
save_with_footer(df_svatky_valid, "vystup_svatky_kveten")

print("Export vikendů")
df_saturday_valid=has_valid_time(df_saturday)
save_with_footer(df_saturday_valid, "vystup_soboty_kveten")
df_sunday_valid=has_valid_time(df_sunday)
save_with_footer(df_sunday_valid, "vystup_nedele_kveten")
df_vikend_valid = pd.concat([df_saturday_valid, df_sunday_valid], ignore_index=False)
save_with_footer(df_vikend_valid, "pdnyv_vikend")

# ===== výpis =====
print(f"\nHotovo.")
print(f"Soubory jsou ve složce: {output_dir}")