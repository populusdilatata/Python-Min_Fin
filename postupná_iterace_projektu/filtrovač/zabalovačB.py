from pathlib import Path
from datetime import datetime, date

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# ==================================================
# NASTAVENÍ
# ==================================================

ROOT_DIR = r"Pro_S"

TARGET_MONTH = 6
TARGET_YEAR = 2026

# ==================================================
# STYLY
# ==================================================

black_fill = PatternFill(
    fill_type="solid",
    fgColor="000000"
)

yellow_font = Font(
    color="FFFF00"
)

# ==================================================
# FUNKCE
# ==================================================

def parse_date(value):
    """Převede obsah buňky na datum."""

    if value is None:
        return None

    if isinstance(value, datetime):
        return value

    if isinstance(value, date):
        return datetime.combine(
            value,
            datetime.min.time()
        )

    if isinstance(value, str):

        value = value.strip()

        if value == "":
            return None

        formats = (
            "%d.%m.%Y",
            "%d.%m.%y",
            "%Y-%m-%d",
            "%d/%m/%Y",
            "%d.%m.%Y %H:%M",
            "%d.%m.%Y %H:%M:%S"
        )

        for fmt in formats:
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                pass

    return None


def color_row(ws, row_number):
    """Obarví celý řádek."""

    for col in range(1, ws.max_column + 1):

        cell = ws.cell(
            row=row_number,
            column=col
        )

        cell.fill = black_fill
        cell.font = yellow_font


# ==================================================
# HLAVNÍ PROGRAM
# ==================================================

for excel_file in Path(ROOT_DIR).rglob("*.xlsx"):

    try:

        print()
        print(f"Zpracovávám: {excel_file}")

        wb = load_workbook(excel_file)

        # druhý a třetí list
        for sheet_idx in (1, 2):

            if sheet_idx >= len(wb.worksheets):
                continue

            ws = wb.worksheets[sheet_idx]

            print(f"  List: {ws.title}")

            for row_num in range(1, ws.max_row + 1):

                # datum je ve sloupci B
                value = ws.cell(
                    row=row_num,
                    column=2
                ).value

                
                print(
                    row,
                    repr(cell.value),
                    type(cell.value)
                        )

                dt = parse_date(value)

                if dt is None:
                    continue

                if (
                    dt.year != TARGET_YEAR
                    or
                    dt.month != TARGET_MONTH
                ):

                    print(
                        f"    Řádek {row_num}: "
                        f"{dt.strftime('%d.%m.%Y')}"
                    )

                    color_row(
                        ws,
                        row_num
                    )

        wb.save(excel_file)

    except Exception as e:

        print()
        print(f"CHYBA: {excel_file}")
        print(e)
def color_row(ws, row_num):

    for cell in ws[row_num]:
        cell.fill = black_fill
        cell.font = yellow_font


    print(
    "Řádek:", row_num,
    "Použitých buněk:", len(ws[row_num]),
    "max_column:", ws.max_column
    )

print()
print("Hotovo.")