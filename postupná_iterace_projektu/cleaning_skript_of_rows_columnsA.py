import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
import os

# ===== nastavení =====
input_file = "pdnyv.xlsx"
LIMIT_PRO_OSCIS=90000
KOD_PRACVMES_ARBITER=901099000
KOD_PRACVMES_FAU=905098000
output_dir = "vystupy"



# vytvoření složky (pokud neexistuje)
os.makedirs(output_dir, exist_ok=True)

# časová přípona pro všechny soubory
time_suffix = datetime.now().strftime("%H_%M")

# ===== načtení =====
df = pd.read_excel(input_file, engine="openpyxl")

# ===== sloupce =====
col_A = df.columns[0]
col_I = "pracvmes"
col_date = "den"

# ===== převody =====
df[col_date] = pd.to_datetime(df[col_date], errors="coerce")

time_cols = ["priche", "odche"]
for col in time_cols:
    if col in df.columns:
        df[col] = pd.to_datetime(df[col], errors="coerce")

# ===== filtr =====
print("Filter")
df = df[(~df[col_I].isin([KOD_PRACVMES_ARBITER, KOD_PRACVMES_FAU])) & (df[col_A] < LIMIT_PRO_OSCIS)]

# ===== květen =====
df_may = df[df[col_date].dt.month == 5]

# ===== svátky =====
df_1_may = df_may[df_may[col_date].dt.day == 1]
df_8_may = df_may[df_may[col_date].dt.day == 8]
df_svatky = pd.concat([df_1_may, df_8_may], ignore_index=True)

# ===== víkendy =====
df_saturday = df_may[df_may[col_date].dt.weekday == 5]
df_sunday = df_may[df_may[col_date].dt.weekday == 6]
df_vikend= pd.concat([df_saturday, df_sunday])

# ===== funkce exportu =====
def save_with_footer(df_input, base_filename):
    df_copy = df_input.copy()

    # formát data
    df_copy[col_date] = df_copy[col_date].dt.strftime("%d.%m.%Y")

    # formát časů
    for col in time_cols:
        if col in df_copy.columns:
            df_copy[col] = df_copy[col].dt.strftime("%H:%M")

    # název souboru s časem
    filename = f"{base_filename}_{time_suffix}.xlsx"
    filepath = os.path.join(output_dir, filename)

    # uložit
    df_copy.to_excel(filepath, index=False)

    # ===== footer =====
    wb = load_workbook(filepath)
    ws = wb.active

    last_row = ws.max_row
    footer_row = last_row + 10

    now_full = datetime.now().strftime("%d.%m.%Y %H:%M:%S")

    ws.cell(row=footer_row, column=1, value=f"Vytvořeno: {now_full}")
    ws.cell(
        row=footer_row + 1,
        column=1,
        value="Softwarové závod sekce Státní tajemník MinFIN"
    )

    wb.save(filepath)

# ===== exporty =====
print("Export")
save_with_footer(df_1_may, "vystup_1_kveten")
save_with_footer(df_8_may, "vystup_8_kveten")
save_with_footer(df_svatky, "pdnyv_vystup_svatky_kveten")

save_with_footer(df_saturday, "vystup_soboty_kveten")
save_with_footer(df_sunday, "vystup_nedele_kveten")
save_with_footer(df_vikend, "pdnyv_vikend.xlsx")

save_with_footer(df, "pdnyv_vystup.xlsx")

# ===== výpis víkendů =====
saturdays = sorted(df_saturday[col_date].dt.date.unique())
sundays = sorted(df_sunday[col_date].dt.date.unique())

print("Soboty v květnu:")
for d in saturdays:
    print(pd.to_datetime(d).strftime("%d.%m.%Y"))

print("\nNeděle v květnu:")
for d in sundays:
    print(pd.to_datetime(d).strftime("%d.%m.%Y"))

print(f"\nHotovo. Soubory jsou ve složce: {output_dir}")