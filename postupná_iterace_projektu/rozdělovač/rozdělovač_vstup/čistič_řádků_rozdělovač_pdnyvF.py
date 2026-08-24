import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from pathlib import Path
from dataclasses import dataclass

# ===== nastavení =====
@dataclass(frozen=True)
class Config:
    limit_oscis: int = 90000
    arbiter: int = 901099000
    fau: int = 905098000
    #===========================
    saturday = 5
    sunday = 6
    min_empty_priche_count = 25
    #===========================
    year = 2026
    month = 6
    zpracovavany_mesic="06_26"
    #=========================== 
    input_file = "vstup_pdnyv/pdnyv_06.xlsx"
    base_output_dir = "porovnavač"
CFG = Config()


@dataclass(frozen=True)
class Columns:
    den: str = "den"
    oscis: str = "oscis"
    pracvmes: str = "pracvmes"
    duvodt: str = "duvodt"

COL = Columns()


HOLIDAYS = {
    1: [1],          # Nový rok
    5: [1, 8],       # Svátek práce, Den vítězství
    7: [5, 6],       # Cyril a Metoděj, Jan Hus
    9: [28],         # Den české státnosti
    10: [28],        # Vznik ČSR
    11: [17],        # Den boje za svobodu
    12: [24, 25, 26] # Vánoce
}


EXCEL_EPOCH = datetime(1899, 12, 30)
# ===== sloupce =====
TIME_COLS = {
    "priche",
    "odche",
    "prichv",
    "odchv",
    "prichsm",
    "odchsm",
    "odprac",
    "odpracz",
    "dfond"
}
#data_pro_rozdělení

CATEGORY_MAP= {
    "dovol": "dovol",
    "puldo": "dovol",

    "odpra": "odpra",
    "služ": "odpra",

    "absen": "absen",
    "nemoc": "nemoc",
    "indiv": "indiv",
    "lékař": "lékař"
}



# ===== složky =====
def create_output_directory() -> tuple[Path, str]:
    
    """
    Vytvoří výstupní adresář pro aktuální běh programu.

    Název adresáře obsahuje datum a čas spuštění,
    což umožňuje uchovávat výsledky jednotlivých běhů odděleně.

    Returns:
        tuple[Path, str]:
            output_dir - cesta k výstupní složce
            time_suffix - čas použitý v názvech souborů
    """
    now = datetime.now()

    run_timestamp = now.strftime("%Y-%m-%d_%H-%M")
    time_suffix = now.strftime("%H_%M")

    base_output_dir = Path("porovnavač")
    base_output_dir.mkdir(exist_ok=True)

    output_dir = base_output_dir / run_timestamp
    output_dir.mkdir(exist_ok=True)

    return output_dir, time_suffix


# ===== načtení =====
def load_data()-> pd.DataFrame:
    """
    Načte vstupní Excel soubor do DataFrame.

    Současně převede sloupec s datem na pandas datetime
    a neplatné hodnoty nahradí NaT.

    Returns:
        pd.DataFrame:
            Načtená a základně připravená data.
    """
    df = pd.read_excel( CFG.input_file, engine="openpyxl")
    df[COL.den] = pd.to_datetime( df[COL.den], errors="coerce")
    return df

# ===== bezpečný převod času =====
def normalize_time_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Převede časové sloupce typu timedelta na datetime.

    Jako základ je použito datum EXCEL_EPOCH,
    ke kterému se přičítá časová složka.

    Args:
        df: Zdrojový DataFrame.

    Returns:
        pd.DataFrame:
            Upravený DataFrame.
    """
    for col in TIME_COLS:

        if col not in df.columns:
            continue

        
        if pd.api.types.is_timedelta64_dtype(df[col]):
            df[col] = EXCEL_EPOCH + df[col]


    return df

# ===== filtr =====
def apply_main_filter(df: pd.DataFrame) -> pd.DataFrame:    
    """
    Aplikuje hlavní filtr záznamů.

    Odstraňuje vybrané organizační jednotky
    a záznamy s OSCIS nad nastaveným limitem.

    Args:
        df: Zdrojový DataFrame.

    Returns:
        pd.DataFrame:
            Filtrovaný DataFrame.
    """
    filter_mask = (~df[COL.pracvmes].isin([
                                                CFG.arbiter,
                                                CFG.fau
                                                ])
                    ) & (
                        df[COL.oscis] < CFG.limit_oscis
                        )

    return df.loc[filter_mask]

# ===== filtr prázdné priche =====
def get_empty_priche(df: pd.DataFrame) -> pd.DataFrame:
   """
    Vybere záznamy bez vyplněného času příchodu.

    Pokud sloupec 'priche' ve vstupních datech
    neexistuje, vrací prázdný DataFrame.

    Args:
        df: Zdrojový DataFrame.

    Returns:
        pd.DataFrame:
            Záznamy s prázdným příchodem.
    """
   if "priche" not in df.columns:
        return pd.DataFrame()
   return df.loc[df["priche"].isna()]


# FUNKCE: filtr na platné časy
def has_valid_time(df_input: pd.DataFrame) -> pd.DataFrame:    
    """
    Odstraní záznamy bez příchodu nebo odchodu.

    Zachová pouze řádky, které obsahují
    vyplněné oba časové údaje.

    Args:
        df_input: Vstupní DataFrame.

    Returns:
        pd.DataFrame:
            Filtrovaná data.
    """
    #Funkce odstrani záznamy, které NEMAJÍ vyplněn sloupeček "priche", nebo "odche"   
    return df_input.dropna(subset=["priche", "odche"])

# ===== APLIKACE EXCEL LOGIKY =====
def filter_frequent_empty_priche(df_prichv_empty: pd.DataFrame) -> pd.DataFrame:
    """
    Vybere zaměstnance s častým výskytem
    prázdného času příchodu.

    Ponechá pouze OSCIS, které se vyskytují
    vícekrát než určuje konfigurace.

    Args:
        df_prichv_empty: Záznamy s prázdným příchodem.

    Returns:
        pd.DataFrame:
            Vyfiltrovaná data.
    """
    if df_prichv_empty.empty:
        return pd.DataFrame()

    counts = (
        df_prichv_empty[COL.oscis]
        .value_counts()
    )

    valid_values = (
        counts[counts > CFG.min_empty_priche_count]
        .index
    )

    return df_prichv_empty.loc[
        df_prichv_empty[COL.oscis]
        .isin(valid_values)
    ]

# definice svátků (měsíc: seznam dní)
def get_holidays(df: pd.DataFrame) -> pd.DataFrame:
    """
    Vyhledá záznamy spadající na státní svátky.

    Svátky jsou definovány ve slovníku HOLIDAYS,
    kde klíčem je měsíc a hodnotou seznam dní.

    Args:
        df: Zdrojový DataFrame.

    Returns:
        pd.DataFrame:
            Záznamy odpovídající svátkům.
    """

    holiday_mask = pd.Series(False, index=df.index)

    for month, days in HOLIDAYS.items():

        holiday_mask |= (
            (df[COL.den].dt.month == month)
            &
            (df[COL.den].dt.day.isin(days))
        )

    return df.loc[holiday_mask]
# definice sobot a neděl
def get_weekends(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Rozdělí víkendové záznamy.

    Vrací samostatně soboty, neděle
    a jejich společné spojení.

    Args:
        df: Zdrojový DataFrame.

    Returns:
        tuple:
            (soboty, neděle, víkend)
    """
    df_saturday = df.loc[df[COL.den].dt.weekday == CFG.saturday]
    df_sunday = df.loc[df[COL.den].dt.weekday == CFG.sunday]

    df_vikend = pd.concat([
        df_saturday,
        df_sunday
    ])

    return (
        df_saturday,
        df_sunday,
        df_vikend
    )
# ===== export =====
def build_filename(name: str) -> str:
    """
    Vytvoří standardizovaný název exportního souboru.

    Všechny exporty používají jednotný formát
    názvu podle zpracovávaného období.

    Args:
        name: Název typu exportu.

    Returns:
        str:
            Název souboru bez přípony.
    """
    return f"pdnyv_vystup_{CFG.zpracovavany_mesic}_{name}"

def save_with_footer(df_input: pd.DataFrame, base_filename: str) -> None:
    """
    Uloží DataFrame do Excel souboru.

    Současně upraví formát datumových a časových
    sloupců tak, aby byly správně zobrazeny
    v Microsoft Excelu.

    Args:
        df_input: DataFrame k exportu.
        base_filename: Základ názvu souboru.
    """
    if df_input.empty:
        print(f"Přeskočeno: {base_filename}")
        return

    print("save_with_footer")
    df_copy = df_input.copy()
    df_copy = df_input.copy()

    df_copy[COL.den] = (
        df_copy[COL.den]
        .dt.strftime("%d.%m.%Y")
    )

    filename = f"{base_filename}_{time_suffix}.xlsx"
    filepath = output_dir / filename

    
    print(df_copy[COL.den].dtype)

    print(
        df_copy[COL.den]
        .head(5)
        .tolist()
        )


    df_copy.to_excel(filepath, index=False)

    wb = load_workbook(filepath)
    ws = wb.active

    # formát času
    for col_idx, col_name in enumerate(df_copy.columns, 1):
        if col_name in TIME_COLS:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).number_format = "HH:MM"
                
    wb.save(filepath)



def export_weekends(df_saturday: pd.DataFrame, df_sunday: pd.DataFrame) -> None:
    
    """
    Exportuje víkendové záznamy.

    Vytváří samostatný export sobot,
    neděl a jejich společný soubor.

    Args:
        df_saturday: Záznamy za soboty.
        df_sunday: Záznamy za neděle.
    """


    print("Export víkendů")
    df_saturday_valid = has_valid_time(df_saturday)

    save_with_footer(
        df_saturday_valid,
        build_filename("soboty")
    )

    df_sunday_valid = has_valid_time(
        df_sunday
    )

    save_with_footer(
        df_sunday_valid,
        build_filename("nedele")
    )

    df_vikend_valid = pd.concat(
        [
            df_saturday_valid,
            df_sunday_valid
        ]
    )

    save_with_footer(
        df_vikend_valid,
        build_filename("vikendy")
    )


def export_categories(df_clean: pd.DataFrame)-> None:
    """
    Rozdělí data podle kategorií důvodů.

    Pro každou nalezenou kategorii vytvoří
    samostatný exportní Excel soubor.

    Args:
        df_clean: DataFrame obsahující sloupec kategorie.
    """
    print("Exporty rozdělených do kategorií")
    for name, df_part in (df_clean.groupby("kategorie")):
        save_with_footer(
            df_part.drop(
                columns=["kategorie"]
            ),
            f"pdnyv_{CFG.zpracovavany_mesic}_FILTER_{name}"
        )
def export_holidays(df_svatky: pd.DataFrame) -> None:
    """
    Exportuje záznamy spadající na státní svátky.

    Před exportem odstraní řádky,
    které nemají vyplněný příchod a odchod.

    Args:
        df_svatky: Záznamy za svátky.
    """

    print("Export svátků")
    save_with_footer(
        has_valid_time(df_svatky),
        build_filename("svatky")
    )


def export_empty_priche(df_prichv_empty_filtered: pd.DataFrame)-> None:
    """
    Exportuje zaměstnance s chybějícím příchodem.

    Do exportu jsou zahrnuti pouze uživatelé,
    kteří překročili nastavený limit výskytů.

    Args:
        df_prichv_empty_filtered:
            Vyfiltrované záznamy.
    """

    print("Export prázdných příchodů")
    save_with_footer(
        df_prichv_empty_filtered,
        build_filename("FILTER_prichv_EMPTY")
    )

def prepare_clean_dataframe(df: pd.DataFrame, df_vikend: pd.DataFrame, df_svatky: pd.DataFrame, df_prichv_empty_filtered: pd.DataFrame) -> pd.DataFrame:

    excluded_idx = (
        set(df_vikend.index)
        | set(df_svatky.index)
        | set(df_prichv_empty_filtered.index)
    )

    df_clean = df.drop(index=excluded_idx).copy()

    df_clean.loc[:, "kategorie"] = (df_clean[COL.duvodt].map(CATEGORY_MAP).fillna("droby"))

    return df_clean

def main():
    """
    Hlavní řídicí funkce programu.

    Zajišťuje načtení dat, aplikaci filtrů,
    přípravu exportních datasetů a vytvoření
    všech výsledných Excel souborů.

    Program začíná i končí právě zde.
    """
    global output_dir
    global time_suffix

    output_dir, time_suffix = (create_output_directory())

    df = load_data()

    df = normalize_time_columns(df)

    df = apply_main_filter(df)
    
    df_prichv_empty = get_empty_priche(df)

    
    df_prichv_empty_filtered = (filter_frequent_empty_priche(df_prichv_empty))


    df_svatky = get_holidays(df)

    (
        df_saturday,
        df_sunday,
        df_vikend
    ) = get_weekends(df)

    print("Exporty")
    # ===== exporty =====
    # nový výstup dle FILTER + COUNTIF
      
    export_empty_priche(df_prichv_empty_filtered)
    df_clean = prepare_clean_dataframe(df, df_vikend, df_svatky, df_prichv_empty_filtered)

    export_holidays(df_svatky)
    export_weekends(df_saturday, df_sunday)

    export_categories(df_clean)
    # ===== výpis =====
    print(f"\nHotovo.")
    print(f"Soubory jsou ve složce: {output_dir}")


if __name__ == "__main__":
    main()
